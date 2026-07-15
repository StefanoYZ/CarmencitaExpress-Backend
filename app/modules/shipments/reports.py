from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.core.business_time import business_now
from app.modules.quotes.service import calculate_quote_for_shipment
from app.modules.shipments.model import Shipment
from app.modules.shipments.repository import list_shipments
from app.modules.sunat.model import ElectronicReceipt


REPORT_HEADERS = [
    "Codigo",
    "Remitente",
    "Destinatario",
    "Peso (kg)",
    "Tarifa (S/)",
    "Estado de pago",
    "Estado",
    "Origen",
    "Destino",
    "Fecha",
]


def build_operational_report_rows(
    db: Session,
    *,
    report_date: date | None = None,
    status: str | None = None,
    search: str | None = None,
) -> list[list[str]]:
    shipments = _filter_shipments(
        list_shipments(db),
        report_date=report_date,
        status=status,
        search=search,
    )
    receipt_by_shipment = {
        receipt.shipment_id: receipt
        for receipt in db.query(ElectronicReceipt)
        .filter(ElectronicReceipt.shipment_id.in_([item.id for item in shipments]))
        .all()
    } if shipments else {}

    rows: list[list[str]] = []
    for shipment in shipments:
        receipt = receipt_by_shipment.get(shipment.id)
        total = receipt.total if receipt is not None else calculate_quote_for_shipment(shipment).total
        rows.append(
            [
                shipment.shipment_code,
                shipment.sender_name,
                shipment.recipient_name,
                f"{shipment.weight_kg:.2f}",
                f"{total:.2f}",
                _payment_status(shipment, receipt),
                shipment.status,
                shipment.origin,
                shipment.destination,
                _format_datetime(shipment.created_at),
            ]
        )
    return rows


def generate_operational_report_pdf(db: Session, **filters) -> bytes:
    rows = build_operational_report_rows(db, **filters)
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Reporte operativo de encomiendas",
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("CARMENCITA EXPRESS CARGO", styles["Title"]),
        Paragraph("Reporte operativo de encomiendas", styles["Heading2"]),
        Paragraph(f"Generado: {business_now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 5 * mm),
    ]
    table_data = [REPORT_HEADERS] + (rows or [["Sin registros"] + [""] * (len(REPORT_HEADERS) - 1)])
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            24 * mm,
            35 * mm,
            35 * mm,
            18 * mm,
            20 * mm,
            25 * mm,
            24 * mm,
            24 * mm,
            27 * mm,
            29 * mm,
        ],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#28A745")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E1D6")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(table)
    document.build(story)
    return buffer.getvalue()


def generate_operational_report_excel(db: Session, **filters) -> bytes:
    """Genera un archivo Excel real (.xlsx / OOXML) con openpyxl."""
    rows = build_operational_report_rows(db, **filters)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Encomiendas"

    worksheet.append(REPORT_HEADERS)
    header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _filter_shipments(
    shipments: list[Shipment],
    *,
    report_date: date | None,
    status: str | None,
    search: str | None,
) -> list[Shipment]:
    normalized_status = status.strip().upper() if status else None
    normalized_search = search.strip().lower() if search else None
    result = []
    for shipment in shipments:
        created_date = shipment.created_at.date() if shipment.created_at else None
        if report_date and created_date != report_date:
            continue
        if normalized_status and shipment.status != normalized_status:
            continue
        if normalized_search:
            searchable = " ".join(
                [
                    shipment.shipment_code or "",
                    shipment.sender_name or "",
                    shipment.recipient_name or "",
                    shipment.origin or "",
                    shipment.destination or "",
                    shipment.description or "",
                ]
            ).lower()
            if normalized_search not in searchable:
                continue
        result.append(shipment)
    return result


def _payment_status(shipment: Shipment, receipt: ElectronicReceipt | None) -> str:
    if shipment.status == "ANULADA":
        return "ANULADO"
    if shipment.status == "PRE_REGISTRADA":
        return "PENDIENTE"
    if receipt is not None:
        return receipt.status or "CONFIRMADO"
    return "CONFIRMADO"


def _format_datetime(value: datetime | None) -> str:
    return value.strftime("%d/%m/%Y %H:%M") if value else "-"
